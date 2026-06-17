from __future__ import annotations

import importlib.util
import io
import json
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


APP_DIR = Path(__file__).resolve().parent
ANALYZER_PATH = APP_DIR / "03_vocab&click.py"
CEFR_PATH = APP_DIR / "merged_cefr.csv"

REQUIRED_COLUMNS = ["ID", "Title", "Normal Ver.", "Easy Ver.", "Difficult Ver."]
VOCAB_KEYS = ["vocab", "normal_vocab", "easy_vocab", "difficult_vocab"]
CEFR_ORDER = ["Pre A1", "A1", "A2", "B1", "B2", "C1", "C2"]


st.set_page_config(
    page_title="Vocab & Click Words",
    page_icon="",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(
    """
    <style>
    .block-container { padding-top: 1.4rem; }
    div[data-testid="stMetricValue"] { font-size: 1.35rem; }
    .small-note { color: #5f6368; font-size: 0.9rem; }
    </style>
    """,
    unsafe_allow_html=True,
)


@st.cache_resource
def load_analyzer():
    spec = importlib.util.spec_from_file_location("vocab_click_analyzer", ANALYZER_PATH)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load analyzer: {ANALYZER_PATH}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


@st.cache_data
def load_cefr_wordlist(path: str, mtime: float) -> pd.DataFrame:
    _ = mtime
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    if "cefr_level" in df.columns:
        order = {level: idx for idx, level in enumerate(CEFR_ORDER)}
        df["cefr_sort"] = df["cefr_level"].map(order).fillna(999)
        df = df.sort_values(["cefr_sort", "word"]).drop(columns=["cefr_sort"])
    return df


def create_template_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Story_Confirmed"
    ws.append(REQUIRED_COLUMNS)
    ws.append([
        "OG0001",
        "Sample Title",
        "#SC01\nNormal version text here.",
        "#SC01\nEasy version text here.",
        "#SC01\nDifficult version text here.",
    ])
    widths = [14, 28, 52, 52, 52]
    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        cell = ws.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def read_story_input(uploaded_file) -> pd.DataFrame:
    return pd.read_excel(uploaded_file, dtype=str)


def validate_story_df(df: pd.DataFrame) -> list[str]:
    missing = [col for col in REQUIRED_COLUMNS if col not in df.columns]
    if missing:
        return [f"필수 컬럼 없음: {', '.join(missing)}"]

    errors: list[str] = []
    for _, row in df.iterrows():
        blanks = [
            col
            for col in ["Normal Ver.", "Easy Ver.", "Difficult Ver."]
            if pd.isna(row[col]) or not str(row[col]).strip()
        ]
        if blanks:
            errors.append(f"{row['ID']} / {row['Title']}: {', '.join(blanks)} 비어 있음")
    return errors


def load_checkpoint(uploaded_file) -> dict:
    if uploaded_file is None:
        return {}
    return json.load(uploaded_file)


def result_is_complete(result: dict | None) -> bool:
    if not result:
        return False
    return all(result.get(key) for key in VOCAB_KEYS)


def build_output_workbook(source_df: pd.DataFrame, results: list[dict], analyzer) -> bytes:
    result_map = {str(result["id"]): result for result in results}
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

    headers = REQUIRED_COLUMNS + [
        "CEFR",
        "CEFR 근거",
        "Lexile",
        "Lexile 근거",
        "Vocab",
        "Normal 어휘",
        "Easy 어휘",
        "Difficult 어휘",
    ]
    widths = [10, 24, 55, 55, 55, 8, 42, 8, 42, 35, 50, 50, 50]
    original_fill = PatternFill("solid", start_color="404040")
    analysis_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 32
    for col_idx, (header, width) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = original_fill if col_idx <= len(REQUIRED_COLUMNS) else analysis_fill
        cell.font = header_font
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    fills = [
        PatternFill("solid", start_color="F2F2F2"),
        PatternFill("solid", start_color="FFFFFF"),
    ]

    for row_idx, (_, source_row) in enumerate(source_df.iterrows(), 2):
        result = result_map.get(str(source_row["ID"]), {})
        row_values = [
            source_row["ID"],
            source_row["Title"],
            source_row["Normal Ver."],
            source_row["Easy Ver."],
            source_row["Difficult Ver."],
            result.get("cefr", ""),
            result.get("cefr_rationale", ""),
            result.get("lexile", ""),
            result.get("lexile_rationale", ""),
            analyzer.vocab_to_cell(result.get("vocab", [])),
            analyzer.vocab_to_cell(result.get("normal_vocab", [])),
            analyzer.vocab_to_cell(result.get("easy_vocab", [])),
            analyzer.vocab_to_cell(result.get("difficult_vocab", [])),
        ]
        fill = fills[row_idx % 2]
        for col_idx, value in enumerate(row_values, 1):
            cell = ws.cell(row_idx, col_idx, "" if pd.isna(value) else value)
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(
                horizontal="center" if col_idx in {1, 6, 8} else "left",
                vertical="top",
                wrap_text=col_idx in {1, 2, 6, 7, 8, 9, 10, 11, 12, 13},
            )
        ws.row_dimensions[row_idx].height = 45

    ws.freeze_panes = "F2"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def checkpoint_to_bytes(results: list[dict]) -> bytes:
    merged = {str(result["id"]): result for result in results}
    return json.dumps(merged, ensure_ascii=False, indent=2).encode("utf-8")


def guide_tab():
    st.markdown("#### 어휘 추출 기준")
    rules = pd.DataFrame(
        [
            {
                "분류": "레벨 초과 어휘",
                "설명": "EVP에서 텍스트 CEFR 밴드보다 같거나 높게 분류된 어휘를 포함합니다.",
                "예시": "A2 텍스트의 age (A2) - 허용\nA2 텍스트의 blossom (B2) - 허용\nA2 텍스트의 all (A1) - 불가",
            },
            {
                "분류": "콘텐츠 특화 어휘",
                "설명": "레벨은 낮아도 해당 주제 배경지식 없이는 모르거나, 해당 스토리 이해에 핵심인 어휘를 포함합니다.",
                "예시": "pea-shooter, pod, moss",
            },
            {
                "분류": "콘텐츠 특화 어휘",
                "설명": "뜻 자체는 쉬워도 비유, 관용 표현, 구동사처럼 문맥에서 특수하게 쓰이는 표현을 포함합니다.",
                "예시": "have, take, give up, break through",
            },
            {
                "분류": "의성어/의태어",
                "설명": "소리나 움직임을 직접 표현하는 단어도 스토리 이해와 클릭 활동에 도움이 되면 포함합니다.",
                "예시": "pop, splash, zoom, crack, chomp",
            },
            {
                "분류": "제외 대상",
                "설명": "고유명사 또는 맥락으로 100% 추론 가능한 어휘는 제외합니다.",
                "예시": "Hans, Christmas Eve",
            },
        ]
    )
    st.dataframe(rules, use_container_width=True, hide_index=True)

    st.markdown("#### 결과 컬럼 의미")
    st.write(
        "- `Normal 어휘`, `Easy 어휘`, `Difficult 어휘`: 각 버전 본문 기준으로 뽑은 학습 어휘입니다.\n"
        "- `Vocab`: Normal 어휘 중 이미지/영상/클릭 활동으로 제시하기 좋은 핵심 단어만 추린 목록입니다.\n"
        "- 동사는 기본형으로 정리하고, 구동사는 목적어 대명사를 빼고 표현 단위로 정리합니다."
    )

    st.divider()
    st.subheader("CEFR 단어 리스트")

    if not CEFR_PATH.exists():
        st.warning("merged_cefr.csv 파일을 찾을 수 없습니다.")
        return

    cefr_df = load_cefr_wordlist(str(CEFR_PATH), CEFR_PATH.stat().st_mtime)
    filter_cols = st.columns([1.2, 1.2, 2])
    with filter_cols[0]:
        levels = st.multiselect(
            "CEFR",
            [level for level in CEFR_ORDER if level in set(cefr_df.get("cefr_level", []))],
            default=[],
        )
    with filter_cols[1]:
        sources = st.multiselect(
            "Source",
            sorted(cefr_df["source"].dropna().unique()) if "source" in cefr_df.columns else [],
            default=[],
        )
    with filter_cols[2]:
        query = st.text_input("단어 검색", placeholder="예: glow, root, cozy")

    filtered = cefr_df.copy()
    if levels and "cefr_level" in filtered.columns:
        filtered = filtered[filtered["cefr_level"].isin(levels)]
    if sources and "source" in filtered.columns:
        filtered = filtered[filtered["source"].isin(sources)]
    if query and "word" in filtered.columns:
        filtered = filtered[
            filtered["word"].str.contains(query.strip(), case=False, na=False)
        ]

    metric_cols = st.columns(3)
    metric_cols[0].metric("전체 단어", f"{len(cefr_df):,}")
    metric_cols[1].metric("현재 결과", f"{len(filtered):,}")
    metric_cols[2].metric(
        "CEFR 단계",
        f"{filtered['cefr_level'].nunique() if 'cefr_level' in filtered else 0}",
    )

    display_columns = [
        column
        for column in [
            "word",
            "cefr_level",
            "source",
            "all_references",
            "additional_references",
        ]
        if column in filtered.columns
    ]
    display_df = filtered[display_columns].rename(
        columns={
            "word": "단어",
            "cefr_level": "주 CEFR",
            "source": "주 출처",
            "all_references": "전체 참고",
            "additional_references": "추가 참고",
        }
    )

    st.dataframe(display_df, use_container_width=True, height=420, hide_index=True)
    st.download_button(
        "필터 결과 CSV 다운로드",
        filtered.to_csv(index=False).encode("utf-8-sig"),
        file_name="filtered_cefr_wordlist.csv",
        mime="text/csv",
    )


def extraction_tab():
    analyzer = load_analyzer()
    st.subheader("어휘 추출")

    st.download_button(
        "업로드용 엑셀 포맷 다운로드",
        create_template_workbook(),
        file_name="Story_Confirmed_Template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    uploaded_story = st.file_uploader(
        "분석할 Story_Confirmed 형식 XLSX 업로드",
        type=["xlsx"],
        accept_multiple_files=False,
    )
    if uploaded_story is None:
        st.info("업로드용 포맷을 내려받아 작성한 뒤 XLSX 파일을 업로드하세요.")
        return

    checkpoint_file = st.file_uploader(
        "이전 분석 결과 JSON 업로드 (선택)",
        type=["json"],
        accept_multiple_files=False,
        help="이전에 다운로드한 analysis_checkpoint.json입니다. 같은 ID의 완료 결과를 재사용해 API 호출 시간과 비용을 줄입니다.",
    )
    with st.expander("checkpoint JSON이 무엇인가요?"):
        st.write(
            "어휘 추출이 끝난 결과를 저장한 재사용 파일입니다. 다음에 같은 스토리를 다시 분석할 때 이 JSON을 업로드하면 "
            "이미 완료된 ID는 건너뛰고 새 ID만 분석할 수 있습니다. 처음 실행하거나 전부 새로 분석하려면 업로드하지 않아도 됩니다."
        )

    try:
        story_df = read_story_input(uploaded_story)
    except Exception as exc:
        st.error(f"입력 파일을 읽지 못했습니다: {exc}")
        return

    errors = validate_story_df(story_df)
    preview_cols = [col for col in REQUIRED_COLUMNS if col in story_df.columns]

    metric_cols = st.columns(4)
    metric_cols[0].metric("스토리", f"{len(story_df):,}")
    metric_cols[1].metric("필수 컬럼", f"{len(preview_cols)}/{len(REQUIRED_COLUMNS)}")
    metric_cols[2].metric("입력 오류", f"{len(errors):,}")
    metric_cols[3].metric("CEFR 파일", "있음" if CEFR_PATH.exists() else "없음")

    with st.expander("입력 미리보기", expanded=True):
        st.dataframe(story_df[preview_cols].head(20), use_container_width=True, hide_index=True)

    if errors:
        st.error("입력 파일을 먼저 수정해야 합니다.")
        st.code("\n".join(errors[:30]), language="text")
        return

    checkpoint = load_checkpoint(checkpoint_file)
    completed_ids = {
        str(sid) for sid, result in checkpoint.items() if result_is_complete(result)
    }

    selection_labels = [
        f"{row['ID']} | {row['Title']}"
        for _, row in story_df.iterrows()
    ]
    selected_labels = st.multiselect(
        "분석 대상",
        selection_labels,
        default=[
            label for label in selection_labels
            if label.split(" | ", 1)[0] not in completed_ids
        ] or selection_labels,
    )
    selected_ids = {label.split(" | ", 1)[0] for label in selected_labels}
    selected_df = story_df[story_df["ID"].astype(str).isin(selected_ids)].copy()

    api_key = st.text_input(
        "Gemini API Key",
        type="password",
        placeholder="분석 실행 시에만 사용됩니다. 앱에 저장되지 않습니다.",
    )
    with st.expander("고급 설정"):
        model_name = st.text_input("모델", value="gemini-2.5-flash")

    st.caption(
        f"checkpoint 재사용 가능 항목: {len(completed_ids):,}개 / 현재 선택: {len(selected_df):,}개"
    )

    run_clicked = st.button(
        "어휘 추출 실행",
        type="primary",
        disabled=selected_df.empty,
        use_container_width=True,
    )
    if run_clicked:
        if not api_key.strip():
            st.error("Gemini API Key가 필요합니다.")
            return

        client = analyzer.genai.Client(api_key=api_key.strip())
        results_by_id: dict[str, dict] = dict(checkpoint)
        progress = st.progress(0)
        log_box = st.empty()

        for idx, (_, row) in enumerate(selected_df.iterrows(), 1):
            sid = str(row["ID"])
            log_box.info(f"{idx}/{len(selected_df)} 분석 중: {sid} / {row['Title']}")
            if result_is_complete(results_by_id.get(sid)):
                progress.progress(idx / len(selected_df))
                continue

            result = analyzer.analyze(client, model_name.strip(), row)
            if not result:
                st.warning(f"{sid} 분석 실패")
                progress.progress(idx / len(selected_df))
                continue
            results_by_id[sid] = result
            progress.progress(idx / len(selected_df))

        selected_results = [
            results_by_id[str(row["ID"])]
            for _, row in story_df.iterrows()
            if str(row["ID"]) in results_by_id
        ]
        output_bytes = build_output_workbook(story_df, selected_results, analyzer)
        checkpoint_bytes = checkpoint_to_bytes(list(results_by_id.values()))

        st.session_state["vocab_output_bytes"] = output_bytes
        st.session_state["vocab_checkpoint_bytes"] = checkpoint_bytes
        st.session_state["vocab_result_count"] = len(selected_results)
        log_box.success(f"완료: {len(selected_results):,}개 결과")

    if "vocab_output_bytes" in st.session_state:
        download_cols = st.columns(2)
        with download_cols[0]:
            st.download_button(
                "Story_Analysis.xlsx 다운로드",
                st.session_state["vocab_output_bytes"],
                file_name="Story_Analysis.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with download_cols[1]:
            st.download_button(
                "checkpoint JSON 다운로드",
                st.session_state["vocab_checkpoint_bytes"],
                file_name="analysis_checkpoint.json",
                mime="application/json",
                use_container_width=True,
            )


st.title("Vocab & Click Words")
tab_guide, tab_extract = st.tabs(["가이드", "어휘 추출"])

with tab_guide:
    guide_tab()

with tab_extract:
    extraction_tab()
