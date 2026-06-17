#!/usr/bin/env python3
"""
Story Level Analyzer
  - CEFR level  : Oxford/Cambridge 기준 (Normal 텍스트 기준)
  - Lexile 점수 : MetaMetrics 기준 (Normal 텍스트 기준)
  - 학습 어휘   : Normal / Easy / Difficult 버전별 각 1세트
결과는 Story_Analysis.xlsx 와 analysis_checkpoint.json 으로 저장됩니다.

Requirements:
  pip install google-genai pandas openpyxl
"""

import os
import sys

def pause_and_exit(code: int = 1):
    """Keep the console open before exiting."""
    print("\n  Press Enter to close...", end="", flush=True)
    input()
    sys.exit(code)

import json
import time
import re
try:
    import pandas as pd
except ImportError:
    print("[ERROR] pandas is not installed.")
    print("  Install: pip install pandas")
    pause_and_exit(1)
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent
os.chdir(BASE_DIR)

# ── 패키지 체크 ──────────────────────────────────────────────────
try:
    from google import genai
    from google.genai import types as genai_types
except ImportError:
    print("[오류] google-genai 패키지가 없습니다.")
    print("  설치: pip install google-genai")
    pause_and_exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("[오류] openpyxl 패키지가 없습니다.")
    print("  설치: pip install openpyxl")
    pause_and_exit(1)

# ── 설정 ─────────────────────────────────────────────────────────
INPUT_FILE      = "Story_Confirmed.xlsx"
OUTPUT_FILE     = "Story_Analysis.xlsx"
CHECKPOINT_FILE = "analysis_checkpoint.json"

# 모델 선택 우선순위 키워드 (이름에 포함된 순서대로 선호)
PREFERRED_KEYWORDS = ["flash", "pro"]

# ── 유틸리티 ──────────────────────────────────────────────────────

def pause_and_exit(code: int = 1):
    """오류 시 창이 바로 닫히지 않도록 Enter 대기 후 종료."""
    print("\n  [Enter]를 눌러 닫으세요...", end="", flush=True)
    input()
    sys.exit(code)


# ── 초기화 ───────────────────────────────────────────────────────

def get_api_key() -> str:
    key = os.environ.get("GEMINI_API_KEY", "").strip()
    if key:
        print("  API 키: 환경변수 GEMINI_API_KEY 사용")
        return key
    print("  Gemini API 키를 입력하세요: ", end="", flush=True)
    return input().strip()


def list_available_models(client) -> list[str]:
    """generateContent를 지원하는 Gemini 모델 목록 반환."""
    candidates = []
    try:
        for m in client.models.list():
            name = getattr(m, "name", "")
            methods = getattr(m, "supported_actions", None) or \
                      getattr(m, "supported_generation_methods", [])
            if "gemini" not in name.lower():
                continue
            # embedding / vision-only 모델 제외
            if any(x in name.lower() for x in ["embed", "aqa"]):
                continue
            methods_str = str(methods).lower()
            if "generatecontent" in methods_str or "generate_content" in methods_str:
                candidates.append(name)
    except Exception as e:
        print(f"  [경고] 모델 목록 조회 실패: {e}")
    return candidates


def pick_best_model(names: list[str]) -> str:
    """flash > pro 순, 버전 번호 높은 것 우선."""
    def score(n):
        n = n.lower()
        # 최신 버전 번호 추출 (2.5 > 2.0 > 1.5 등)
        ver = 0.0
        m = re.search(r"gemini-(\d+\.\d+)", n)
        if m:
            ver = float(m.group(1))
        flash = 1 if "flash" in n else 0
        # preview/exp 제품은 약간 낮게
        preview = -0.1 if any(x in n for x in ["preview", "exp", "latest"]) else 0
        return (ver, flash, preview)
    return max(names, key=score)


def setup_model(api_key: str):
    client = genai.Client(api_key=api_key)

    print("  사용 가능한 모델 조회 중...", flush=True)
    available = list_available_models(client)

    if not available:
        print("  [오류] generateContent 지원 모델을 찾을 수 없습니다.")
        print("  API 키 권한 또는 네트워크를 확인해 주세요.")
        pause_and_exit(1)

    chosen = pick_best_model(available)
    # 모델명이 "models/..." 형식이면 앞 부분 제거
    short = chosen.replace("models/", "")

    print(f"  사용 모델: {short}  (전체 {len(available)}개 중 선택)")
    return client, short


# ── 데이터 ───────────────────────────────────────────────────────

def load_stories() -> pd.DataFrame:
    if not Path(INPUT_FILE).exists():
        print(f"[오류] {INPUT_FILE} 파일을 찾을 수 없습니다.")
        pause_and_exit(1)
    df = pd.read_excel(INPUT_FILE)
    required = ["ID", "Title", "Normal Ver.", "Easy Ver.", "Difficult Ver."]
    missing = [c for c in required if c not in df.columns]
    if missing:
        print(f"[오류] {INPUT_FILE}에 필수 컬럼이 없습니다: {', '.join(missing)}")
        pause_and_exit(1)

    blank_rows = []
    for _, row in df.iterrows():
        missing_versions = [
            col for col in ["Normal Ver.", "Easy Ver.", "Difficult Ver."]
            if pd.isna(row[col]) or not str(row[col]).strip()
        ]
        if missing_versions:
            blank_rows.append((row["ID"], row["Title"], ", ".join(missing_versions)))

    if blank_rows:
        print(f"[오류] {INPUT_FILE}에 비어 있는 본문 컬럼이 있습니다.")
        print("Easy/Difficult 어휘는 해당 버전 본문이 있어야 추출됩니다.")
        print()
        for sid, title, cols in blank_rows[:20]:
            print(f"  - {sid} / {title}: {cols}")
        if len(blank_rows) > 20:
            print(f"  ... 외 {len(blank_rows) - 20}개")
        pause_and_exit(1)

    return df


def load_checkpoint() -> dict:
    if Path(CHECKPOINT_FILE).exists():
        with open(CHECKPOINT_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_checkpoint(data: dict):
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── 사용자 인터페이스 ─────────────────────────────────────────────

def show_list(df: pd.DataFrame, checkpoint: dict):
    done = sum(1 for _, r in df.iterrows() if r["ID"] in checkpoint)
    print(f"\n{'-'*62}")
    print(f"  {'ID':<10} {'제목':<36} 상태")
    print(f"{'-'*62}")
    for _, row in df.iterrows():
        status = "[완료]" if row["ID"] in checkpoint else "[대기]"
        title  = str(row["Title"])[:34]
        print(f"  {row['ID']:<10} {title:<36} {status}")
    print(f"{'-'*62}")
    print(f"  전체 {len(df)}개 | 완료 {done}개 | 미완료 {len(df)-done}개\n")


def select_stories(df: pd.DataFrame, checkpoint: dict) -> pd.DataFrame:
    show_list(df, checkpoint)
    print("  선택 방법:")
    print("    all          → 전체 분석")
    print("    new          → 미완료 스토리만")
    print("    CS0001       → 특정 스토리 1개")
    print("    CS0001,CS0003 → 복수 선택 (쉼표로 구분)")
    print("    redo CS0001  → 완료된 스토리 재분석")
    print("    q            → 종료")
    choice = input("\n  선택: ").strip()

    if choice.lower() == "q":
        print("  종료합니다.")
        sys.exit(0)
    if choice.lower() == "all":
        return df
    if choice.lower() == "new":
        pending = df[~df["ID"].isin(checkpoint.keys())]
        if pending.empty:
            print("  모든 스토리가 이미 분석되었습니다.")
            print("  재분석하려면 'redo ID' 또는 'all' 을 사용하세요.")
        return pending
    if choice.lower().startswith("redo "):
        rid = choice[5:].strip().upper()
        result = df[df["ID"] == rid]
        if result.empty:
            print(f"  ID '{rid}' 를 찾을 수 없습니다.")
        return result
    # 쉼표 구분 ID 목록
    ids = [x.strip().upper() for x in choice.split(",") if x.strip()]
    result = df[df["ID"].isin(ids)]
    if result.empty:
        print(f"  입력한 ID를 찾을 수 없습니다: {ids}")
    return result


# ── Gemini 분석 ───────────────────────────────────────────────────

def clean_text(raw) -> str:
    """장면 마커(#SC01 등) 제거 후 텍스트 반환"""
    if pd.isna(raw):
        return ""
    return re.sub(r"#SC\d+\s*\n?", "", str(raw)).strip()


def build_prompt(title: str, normal: str, easy: str, difficult: str) -> str:
    return f"""You are a certified literacy assessment expert with deep knowledge of:
- CEFR framework (Oxford/Cambridge standard descriptors for vocabulary range,
  grammatical accuracy and range, coherence, discourse markers)
- Lexile Framework for Reading by MetaMetrics (mean sentence length + word
  frequency/semantic difficulty → Lexile measure)

Analyze the three versions of this children's story and return a single JSON object
with exactly these fields:

"cefr"
  CEFR level for the NORMAL version ONLY.
  Choose one: A1, A2, B1, B2, C1, C2
  Oxford/Cambridge criteria:
    A1: only the most basic vocabulary, present tense, very short simple sentences
    A2: basic everyday vocabulary, simple past, simple compound sentences
    B1: moderate vocabulary range, several tenses, some complex sentences
    B2: broad vocabulary, all tenses, complex grammatical structures
  For a children's story, the typical range is A1–B1.

"cefr_rationale"
  1–2 sentences with specific textual evidence from the Normal version
  (e.g. average sentence length, notable grammar patterns, vocabulary level).

"lexile"
  Lexile estimate for the NORMAL version ONLY using MetaMetrics format.
  Calculation basis: mean sentence length AND word semantic difficulty (frequency).
  Examples: "BR200L" (below 0), "300L", "550L", "750L"
  Reference ranges:
    Pre-K / early readers: 50–300L
    Grade 1–2 readers: 200–500L
    Grade 3–4 readers: 400–700L
    Grade 5–6 readers: 600–850L

"lexile_rationale"
  1 sentence explaining the estimate (avg sentence length + vocabulary difficulty).

"normal_vocab"
  A plain list of vocabulary items from the NORMAL version that should be
  flagged for learners reading at the NORMAL CEFR level. Include ALL words
  that meet EITHER of the two criteria below. No fixed count.

  SELECTION CRITERIA (include a word if it meets criterion A OR B):
    A. EVP LEVEL AT OR ABOVE: The word is classified at the SAME CEFR band as
       the text's level OR at a higher band, according to the Cambridge English
       Vocabulary Profile (EVP).
       For example, if the text is A2, include any word the EVP classifies as
       A2, B1, B2, C1, or C2.
    B. CONTEXT-OPAQUE CONTENT WORD: The word is a content word (noun, verb,
       adjective, adverb) that cannot be reliably inferred from context alone,
       even if it falls within the CEFR band. This covers:
         - Domain/topic-specific words (e.g. "pea-shooter", "pod", "moss")
         - Figurative or idiomatic uses not transparent from literal meaning
         - Words whose meaning is critical to the plot but not signalled by
           surrounding sentences

  EXCLUSIONS (do NOT include any of the following):
    - Proper nouns (names of people, places)
    - Words whose meaning is fully and unambiguously recoverable from
      immediate context without any vocabulary knowledge
    - Onomatopoeia: words used purely to represent a sound effect rather than
      as a content word. Judge by the word's function in that specific sentence.
        * Exclude when used as a sound effect: "Pop!", "Splash!", "Zoom!",
          "Zap!", "CRACK!" (written as an exclamation or isolated interjection)
        * Include when used as a genuine content word in the same or another
          sentence: e.g. "He heard a crack in the wall" (noun) or
          "The ice cracked" (verb) → include as "crack"

  FORM RULES (STRICTLY follow):
    1. Verbs: BASE FORM only (e.g. "fell" → "fall", "blew" → "blow")
    2. Adjectives / nouns / adverbs: form as it appears in the text
       (e.g. "proud", "pea-shooter", "softly")
    3. Phrasal verbs: full phrase WITHOUT object pronoun
       (e.g. "turn it off" → "turn off", "picked them up" → "pick up")
    4. No definitions. Words only.
  Format: ["word1", "word2", "phrasal verb", ...]

"vocab"
  STEP 1 — source: take the list you just produced in "normal_vocab". That is
           your ONLY source. Do NOT go back to the story text. Do NOT add any
           word that does not already appear in normal_vocab.
  STEP 2 — filter: from that normal_vocab list, keep a word ONLY IF it satisfies
           BOTH of the following conditions at the same time:
             a) IMAGEABLE — the word can be depicted as a single clear image or
                short video clip (concrete nouns, dynamic action verbs, vivid
                sensory adjectives). Discard anything abstract or functional.
             b) STORY-CENTRAL & HINT — the word is essential to the story (key
                character trait, object, action, setting, or emotion), AND the
                resulting set of words, heard in sequence, gives a "story trailer"
                feel — key characters, objects and events should be sensed.
  STEP 3 — count (based on the CEFR of the NORMAL version):
             A1 or A2 → 5–8 words   |   B1 or B2 → 6–10 words

  Every word in vocab MUST already be present in normal_vocab.
  No definitions. Words only. Same form as in normal_vocab.
  Format: ["word1", "word2", ...]

"easy_vocab"
  Same SELECTION CRITERIA, EXCLUSIONS, and FORM RULES as normal_vocab,
  applied to the EASY version at its own CEFR level.
  Format: ["word1", "word2", ...]

"difficult_vocab"
  Same SELECTION CRITERIA, EXCLUSIONS, and FORM RULES as normal_vocab,
  applied to the DIFFICULT version at its own CEFR level.
  Format: ["word1", "word2", ...]

────────────────────────────────────────────
STORY TITLE: {title}

[NORMAL VERSION]
{normal}

[EASY VERSION]
{easy}

[DIFFICULT VERSION]
{difficult}
────────────────────────────────────────────

IMPORTANT: Return ONLY a valid JSON object. No markdown fences, no extra text."""


def call_gemini(client, model_name: str, prompt: str, max_retries: int = 3) -> dict | None:
    for attempt in range(1, max_retries + 1):
        try:
            cfg = genai_types.GenerateContentConfig(
                temperature=0.15,
                response_mime_type="application/json",
            )
            full_name = model_name if model_name.startswith("models/") else f"models/{model_name}"
            response = client.models.generate_content(
                model=full_name,
                contents=prompt,
                config=cfg,
            )
            text = response.text.strip()

            # 혹시 markdown 코드 블록이 붙어 있으면 제거
            text = re.sub(r"^```(?:json)?\s*", "", text)
            text = re.sub(r"\s*```$", "", text)
            text = text.strip()

            return json.loads(text)

        except json.JSONDecodeError as e:
            print(f"    JSON 파싱 오류 (시도 {attempt}/{max_retries}): {e}")
        except Exception as e:
            print(f"    API 오류 (시도 {attempt}/{max_retries}): {type(e).__name__}: {e}")

        if attempt < max_retries:
            wait = 4 * attempt
            print(f"    {wait}초 후 재시도...")
            time.sleep(wait)

    return None


def analyze(client, model_name: str, row: pd.Series) -> dict | None:
    normal    = clean_text(row["Normal Ver."])
    easy      = clean_text(row["Easy Ver."])
    difficult = clean_text(row["Difficult Ver."])
    prompt    = build_prompt(str(row["Title"]), normal, easy, difficult)
    result    = call_gemini(client, model_name, prompt)
    if result:
        result["id"]    = row["ID"]
        result["title"] = str(row["Title"])
    return result


# ── Excel 출력 ────────────────────────────────────────────────────

def vocab_to_cell(vocab) -> str:
    """문자열 리스트 또는 구버전 dict 리스트 모두 처리 → 콤마 구분 단어 문자열."""
    if not vocab:
        return "-"
    words = []
    for v in vocab:
        if isinstance(v, str):
            words.append(v.strip())
        elif isinstance(v, dict):
            # 구버전 checkpoint 호환: {"word": "...", "definition": "..."}
            words.append(v.get("word", "").strip())
    return ", ".join(w for w in words if w)


def save_workbook_safely(wb, output_file: str) -> str:
    output_path = Path(output_file)
    try:
        wb.save(output_path)
        return str(output_path)
    except PermissionError:
        alt_path = output_path.with_name(
            f"{output_path.stem}_{time.strftime('%Y%m%d_%H%M%S')}{output_path.suffix}"
        )
        wb.save(alt_path)
        print()
        print(f"[경고] {output_path.name} 파일이 열려 있어 덮어쓸 수 없습니다.")
        print(f"       대신 {alt_path.name} 파일로 저장했습니다.")
        return str(alt_path)


def write_excel(results: list):
    # 원본 데이터 로드
    src_df = pd.read_excel(INPUT_FILE)
    # 분석 결과를 ID 기준 dict로 변환
    result_map = {r["id"]: r for r in results}

    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis"

    # ── 컬럼 정의 ──────────────────────────────────────────────────
    # 원본 5개 + 분석 8개 (Vocab 추가)
    ORIG_HEADERS  = ["ID", "Title", "Normal Ver.", "Easy Ver.", "Difficult Ver."]
    ANAL_HEADERS  = ["CEFR", "CEFR 근거", "Lexile", "Lexile 근거",
                     "Vocab", "Normal 어휘", "Easy 어휘", "Difficult 어휘"]
    ALL_HEADERS   = ORIG_HEADERS + ANAL_HEADERS

    # 열 너비: 원본(ID, Title, 본문×3) + 분석
    COL_WIDTHS = [10, 24, 55, 55, 55,            # 원본
                  8,  42,  8,  42, 35, 50, 50, 50]  # 분석

    # ── 헤더 스타일 ────────────────────────────────────────────────
    # 원본 헤더: 짙은 회색
    ORIG_FILL = PatternFill("solid", start_color="404040")
    # 분석 헤더: 짙은 남색
    ANAL_FILL = PatternFill("solid", start_color="1F3864")
    H_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 32

    for c, (h, w) in enumerate(zip(ALL_HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(1, c, h)
        cell.fill  = ORIG_FILL if c <= len(ORIG_HEADERS) else ANAL_FILL
        cell.font  = H_FONT
        cell.alignment = H_ALIGN
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── 데이터 행 ──────────────────────────────────────────────────
    FILLS = [
        PatternFill("solid", start_color="F2F2F2"),
        PatternFill("solid", start_color="FFFFFF"),
    ]

    for i, (_, src_row) in enumerate(src_df.iterrows()):
        ri   = i + 2
        fill = FILLS[i % 2]
        sid  = src_row["ID"]
        r    = result_map.get(sid, {})

        kv = vocab_to_cell(r.get("vocab",          []))
        nv = vocab_to_cell(r.get("normal_vocab",   []))
        ev = vocab_to_cell(r.get("easy_vocab",     []))
        dv = vocab_to_cell(r.get("difficult_vocab",[]))

        row_data = [
            # 원본
            src_row["ID"],
            src_row["Title"],
            src_row["Normal Ver."],
            src_row["Easy Ver."],
            src_row["Difficult Ver."],
            # 분석
            r.get("cefr", ""),
            r.get("cefr_rationale", ""),
            r.get("lexile", ""),
            r.get("lexile_rationale", ""),
            kv, nv, ev, dv,
        ]

        # 텍스트 열(C~E)은 wrap 없이, 분석 열은 wrap
        WRAP_COLS   = {1, 2, 6, 7, 8, 9, 10, 11, 12, 13}  # wrap=True
        CENTER_COLS = {1, 6, 8}                             # 가운데 정렬

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(ri, c, val if not pd.isna(val) else "")
            cell.fill = fill
            cell.font = Font(name="Arial", size=10)
            wrap = c in WRAP_COLS
            halign = "center" if c in CENTER_COLS else "left"
            cell.alignment = Alignment(
                horizontal=halign, vertical="top", wrap_text=wrap
            )

        ws.row_dimensions[ri].height = 45  # 본문 일부 표시에 적합한 고정 높이

    ws.freeze_panes = "F2"   # 원본 컬럼 고정, 분석 컬럼부터 스크롤
    return save_workbook_safely(wb, OUTPUT_FILE)


# ── 메인 ─────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 62)
    print("  Story Level Analyzer")
    print("  CEFR (Oxford/Cambridge) / Lexile (MetaMetrics) / 학습어휘")
    print("=" * 62 + "\n")

    df = load_stories()
    checkpoint = load_checkpoint()

    api_key = get_api_key()
    if not api_key:
        print("[오류] API 키가 필요합니다.")
        pause_and_exit(1)

    print("\n  Gemini 연결 중...")
    client, model_name = setup_model(api_key)

    while True:
        selected = select_stories(df, checkpoint)
        if not selected.empty:
            break
        print("  다시 선택해 주세요.\n")

    n = len(selected)
    print(f"\n  {n}개 스토리를 분석합니다...\n")

    for idx, (_, row) in enumerate(selected.iterrows(), 1):
        sid   = row["ID"]
        title = str(row["Title"])
        print(f"  [{idx}/{n}] {sid} - {title}")

        result = analyze(client, model_name, row)

        if result:
            checkpoint[sid] = result
            save_checkpoint(checkpoint)
            kv = len(result.get("vocab",          []))
            nv = len(result.get("normal_vocab",   []))
            ev = len(result.get("easy_vocab",     []))
            dv = len(result.get("difficult_vocab",[]))
            print(
                f"        CEFR: {result.get('cefr', '?')}"
                f"  |  Lexile: {result.get('lexile', '?')}"
                f"  |  Vocab: {kv}개"
                f"  |  어휘: {nv}/{ev}/{dv} 개 (Normal/Easy/Difficult)"
            )
        else:
            print(f"        [실패] 분석 실패 - 건너뜀")

        if idx < n:
            time.sleep(1.5)  # API 속도 제한 여유

    # 원본 순서대로 전체 결과 취합 후 저장
    all_results = [
        checkpoint[r["ID"]]
        for _, r in df.iterrows()
        if r["ID"] in checkpoint
    ]

    if all_results:
        saved_output = write_excel(all_results)
        print(f"\n  [저장완료] {len(all_results)}개 결과 -> {saved_output}")
    else:
        print("\n  저장할 결과가 없습니다.")

    print("\n  완료!\n")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  중단되었습니다.")
    except Exception as e:
        import traceback
        print("\n" + "="*62)
        print("[예상치 못한 오류]")
        traceback.print_exc()
        print("="*62)
        pause_and_exit(1)
    print("  [Enter]를 눌러 닫으세요...", end="", flush=True)
    input()
